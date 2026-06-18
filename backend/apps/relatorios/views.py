"""
Dashboard financeiro e gerador de relatórios com exportação Excel/CSV.

Todos os endpoints são GET e requerem autenticação JWT.
Nenhum destes endpoints altera dados.
"""
import calendar
import csv
import io
from datetime import date

from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth, ExtractMonth
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.financeiro.titulos_pagar.models import TituloPagar
from apps.financeiro.titulos_receber.models import TituloReceber

MONTH_LABELS = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

TIPO_RECEBER_LABELS = {
    TituloReceber.TIPO_MENSALIDADE: 'Mensalidade',
    TituloReceber.TIPO_VOO: 'Cobrança de Voo',
    TituloReceber.TIPO_HORAS_PRE_PAGAS: 'Recarga de Carteira',
    TituloReceber.TIPO_SERVICO: 'Serviço',
    TituloReceber.TIPO_OUTROS: 'Outros',
}

TIPO_PAGAR_LABELS = {
    TituloPagar.TIPO_FORNECEDOR: 'Fornecedor',
    TituloPagar.TIPO_FOLHA: 'Folha de Pagamento',
    TituloPagar.TIPO_CONTA_FIXA: 'Conta Fixa',
    TituloPagar.TIPO_OUTROS: 'Outros',
}

# tipos de TituloReceber excluídos dos indicadores financeiros
_EXCLUIR_RECEBER: list = []  # recargas de carteira agora incluídas nos gráficos


def _parse_mes(value: str) -> date:
    """'YYYY-MM' → primeiro dia do mês."""
    year, month = map(int, value.split('-'))
    return date(year, month, 1)


def _fim_mes(value: str) -> date:
    """'YYYY-MM' → último dia do mês."""
    year, month = map(int, value.split('-'))
    _, last = calendar.monthrange(year, month)
    return date(year, month, last)


class DashboardResumoView(APIView):
    """
    GET /api/v1/dashboard/resumo/

    Retorna 7 indicadores financeiros para os cards do painel.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        hoje = timezone.now().date()
        inicio_mes = hoje.replace(day=1)

        # Títulos a receber em aberto (excluindo carteira). Inclui remessa_criada
        # (boleto registrado em remessa CNAB, mas ainda não pago) — continua sendo
        # um recebível em aberto/vencido.
        qs_receber_aberto = (
            TituloReceber.objects
            .exclude(status=TituloReceber.STATUS_BAIXADO)
            .exclude(tipo__in=_EXCLUIR_RECEBER)
        )
        # saldo_devedor é property — precisamos iterar
        total_receber = sum(float(t.saldo_devedor) for t in qs_receber_aberto)
        vencidos_receber = sum(
            float(t.saldo_devedor)
            for t in qs_receber_aberto.filter(data_vencimento__lt=hoje)
        )

        # Títulos a pagar em aberto
        agg_pagar = TituloPagar.objects.filter(
            status=TituloPagar.STATUS_ABERTO
        ).aggregate(total=Sum('valor'))
        total_pagar = float(agg_pagar['total'] or 0)

        agg_venc_pagar = TituloPagar.objects.filter(
            status=TituloPagar.STATUS_ABERTO,
            data_vencimento__lt=hoje,
        ).aggregate(total=Sum('valor'))
        vencidos_pagar = float(agg_venc_pagar['total'] or 0)

        # Recebidos no mês corrente
        agg_rec_mes = (
            TituloReceber.objects
            .filter(status=TituloReceber.STATUS_BAIXADO, data_pagamento__gte=inicio_mes)
            .exclude(tipo__in=_EXCLUIR_RECEBER)
            .aggregate(total=Sum('valor_pago'))
        )
        recebidos_mes = float(agg_rec_mes['total'] or 0)

        # Pagos no mês corrente
        agg_pag_mes = TituloPagar.objects.filter(
            status=TituloPagar.STATUS_BAIXADO,
            data_pagamento__gte=inicio_mes,
        ).aggregate(total=Sum('valor_pago'))
        pagos_mes = float(agg_pag_mes['total'] or 0)

        return Response({
            'total_receber': total_receber,
            'total_pagar': total_pagar,
            'vencidos_receber': vencidos_receber,
            'vencidos_pagar': vencidos_pagar,
            'recebidos_mes': recebidos_mes,
            'pagos_mes': pagos_mes,
            'saldo_mes': recebidos_mes - pagos_mes,
        })


class DashboardVencidosPorMesView(APIView):
    """
    GET /api/v1/dashboard/vencidos-por-mes/

    Query params:
        tipo        receber | pagar  (default: receber)
        metrica     quantidade | valor  (default: valor)
        data_inicio YYYY-MM  (default: 12 meses atrás)
        data_fim    YYYY-MM  (default: mês atual)
        categoria   tipo do título (opcional)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        hoje = timezone.now().date()
        tipo = request.query_params.get('tipo', 'receber')
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        categoria = request.query_params.get('categoria', '')

        # Período padrão: últimos 12 meses
        if data_inicio_str:
            inicio = _parse_mes(data_inicio_str)
        else:
            m = hoje.month - 11
            y = hoje.year + (m - 1) // 12
            m = ((m - 1) % 12) + 1
            inicio = date(y, m, 1)

        fim = _fim_mes(data_fim_str) if data_fim_str else hoje

        if tipo == 'pagar':
            qs = TituloPagar.objects.filter(
                status=TituloPagar.STATUS_ABERTO,
                data_vencimento__lt=hoje,
                data_vencimento__gte=inicio,
                data_vencimento__lte=fim,
            )
            if categoria:
                qs = qs.filter(tipo=categoria)
            grouped = (
                qs.annotate(mes=TruncMonth('data_vencimento'))
                .values('mes')
                .annotate(quantidade=Count('id'), valor=Sum('valor'))
                .order_by('mes')
            )
        else:
            qs = (
                TituloReceber.objects
                .exclude(status=TituloReceber.STATUS_BAIXADO)
                .filter(
                    data_vencimento__lt=hoje,
                    data_vencimento__gte=inicio,
                    data_vencimento__lte=fim,
                )
                .exclude(tipo__in=_EXCLUIR_RECEBER)
            )
            if categoria:
                qs = qs.filter(tipo=categoria)
            grouped = (
                qs.annotate(mes=TruncMonth('data_vencimento'))
                .values('mes')
                .annotate(quantidade=Count('id'), valor=Sum('valor_original'))
                .order_by('mes')
            )

        resultado = []
        for item in grouped:
            mes_date = item['mes']
            resultado.append({
                'mes': f"{mes_date.year}-{mes_date.month:02d}",
                'label': f"{MONTH_LABELS[mes_date.month - 1]}/{str(mes_date.year)[2:]}",
                'quantidade': item['quantidade'],
                'valor': float(item['valor'] or 0),
            })

        return Response(resultado)


class DashboardEntradasPorGrupoView(APIView):
    """
    GET /api/v1/dashboard/entradas-por-grupo/

    Query params:
        data_inicio YYYY-MM-DD  (default: primeiro dia do mês atual)
        data_fim    YYYY-MM-DD  (default: hoje)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        hoje = timezone.now().date()
        data_inicio = request.query_params.get('data_inicio', str(hoje.replace(day=1)))
        data_fim = request.query_params.get('data_fim', str(hoje))

        grouped = (
            TituloReceber.objects
            .filter(
                status=TituloReceber.STATUS_BAIXADO,
                data_pagamento__gte=data_inicio,
                data_pagamento__lte=data_fim,
            )
            .exclude(tipo__in=_EXCLUIR_RECEBER)
            .values('tipo')
            .annotate(valor=Sum('valor_pago'))
            .order_by('-valor')
        )

        total = sum(float(g['valor'] or 0) for g in grouped)

        resultado = []
        for item in grouped:
            v = float(item['valor'] or 0)
            resultado.append({
                'grupo': TIPO_RECEBER_LABELS.get(item['tipo'], item['tipo']),
                'tipo': item['tipo'],
                'valor': v,
                'percentual': round((v / total * 100) if total > 0 else 0, 1),
            })

        return Response(resultado)


class DashboardHistoricoAnualView(APIView):
    """
    GET /api/v1/dashboard/historico-anual/

    Query params:
        ano          (default: ano atual)
        tipo_pagar   tipos separados por vírgula (ex: fornecedor,conta_fixa)
        tipo_receber tipos separados por vírgula (ex: mensalidade,voo)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        hoje = timezone.now().date()
        ano = int(request.query_params.get('ano', hoje.year))
        tipo_pagar_str = request.query_params.get('tipo_pagar', '')
        tipo_receber_str = request.query_params.get('tipo_receber', '')

        meses = [
            {'mes': m, 'label': MONTH_LABELS[m - 1], 'entradas': 0.0, 'saidas': 0.0, 'saldo': 0.0}
            for m in range(1, 13)
        ]

        # Entradas: títulos receber baixados no ano
        qs_receber = (
            TituloReceber.objects
            .filter(status=TituloReceber.STATUS_BAIXADO, data_pagamento__year=ano)
            .exclude(tipo__in=_EXCLUIR_RECEBER)
        )
        if tipo_receber_str:
            tipos = [t.strip() for t in tipo_receber_str.split(',') if t.strip()]
            qs_receber = qs_receber.filter(tipo__in=tipos)

        for row in (
            qs_receber
            .annotate(m=ExtractMonth('data_pagamento'))
            .values('m')
            .annotate(total=Sum('valor_pago'))
        ):
            meses[row['m'] - 1]['entradas'] = float(row['total'] or 0)

        # Saídas: títulos pagar baixados no ano
        qs_pagar = TituloPagar.objects.filter(
            status=TituloPagar.STATUS_BAIXADO,
            data_pagamento__year=ano,
        )
        if tipo_pagar_str:
            tipos = [t.strip() for t in tipo_pagar_str.split(',') if t.strip()]
            qs_pagar = qs_pagar.filter(tipo__in=tipos)

        for row in (
            qs_pagar
            .annotate(m=ExtractMonth('data_pagamento'))
            .values('m')
            .annotate(total=Sum('valor_pago'))
        ):
            meses[row['m'] - 1]['saidas'] = float(row['total'] or 0)

        for m in meses:
            m['saldo'] = round(m['entradas'] - m['saidas'], 2)

        return Response(meses)


# =============================================================================
# Gerador de Relatórios / Exportação
# =============================================================================

class RelatorioMetadadosView(APIView):
    """
    GET /api/v1/relatorios/metadados/

    Retorna os campos disponíveis, tipos e status para montar o formulário
    do gerador de relatórios no frontend.
    """
    permission_classes = [IsAuthenticated]

    # Campos com tipo inteiro (não monetários) — usados pelo frontend para evitar formatação de moeda
    _CAMPOS_INTEIROS = {'num_parcela', 'total_parcelas'}

    def get(self, request):
        return Response({
            'campos': {
                'receber': {
                    'participante_nome': 'Participante',
                    'tipo': 'Categoria',
                    'descricao': 'Descrição',
                    'data_emissao': 'Data de Emissão',
                    'data_vencimento': 'Vencimento',
                    'data_pagamento': 'Data de Pagamento',
                    'valor_original': 'Valor (R$)',
                    'valor_pago': 'Valor Pago (R$)',
                    'multa': 'Multa (R$)',
                    'saldo_devedor': 'Saldo Devedor (R$)',
                    'status': 'Situação',
                    'num_parcela': 'Parcela',
                    'total_parcelas': 'Total Parcelas',
                },
                'pagar': {
                    'favorecido_nome': 'Favorecido',
                    'tipo': 'Tipo',
                    'descricao': 'Descrição',
                    'data_emissao': 'Data de Emissão',
                    'data_vencimento': 'Vencimento',
                    'data_pagamento': 'Data de Pagamento',
                    'valor': 'Valor (R$)',
                    'valor_pago': 'Valor Pago (R$)',
                    'status': 'Situação',
                    'num_parcela': 'Parcela',
                    'total_parcelas': 'Total Parcelas',
                },
            },
            'campos_inteiros': list(self._CAMPOS_INTEIROS),
            'tipos': {
                'receber': [
                    {'value': k, 'label': v}
                    for k, v in TituloReceber.TIPO_CHOICES
                ],
                'pagar': [
                    {'value': k, 'label': v}
                    for k, v in TituloPagar.TIPO_CHOICES
                ],
            },
            'status': {
                'receber': [
                    {'value': k, 'label': v} for k, v in TituloReceber.STATUS_CHOICES
                ],
                'pagar': [
                    {'value': k, 'label': v} for k, v in TituloPagar.STATUS_CHOICES
                ],
            },
            'data_fields': {
                'receber': [
                    {'value': 'data_vencimento', 'label': 'Vencimento'},
                    {'value': 'data_emissao', 'label': 'Emissão'},
                    {'value': 'data_pagamento', 'label': 'Pagamento'},
                ],
                'pagar': [
                    {'value': 'data_vencimento', 'label': 'Vencimento'},
                    {'value': 'data_emissao', 'label': 'Emissão'},
                    {'value': 'data_pagamento', 'label': 'Pagamento'},
                ],
            },
            'ordenar_por': {
                'receber': [
                    {'value': 'data_vencimento', 'label': 'Vencimento'},
                    {'value': 'data_emissao', 'label': 'Emissão'},
                    {'value': 'data_pagamento', 'label': 'Pagamento'},
                    {'value': 'valor_original', 'label': 'Valor'},
                    {'value': 'valor_pago', 'label': 'Valor Pago'},
                    {'value': 'status', 'label': 'Situação'},
                    {'value': 'tipo', 'label': 'Categoria'},
                ],
                'pagar': [
                    {'value': 'data_vencimento', 'label': 'Vencimento'},
                    {'value': 'data_emissao', 'label': 'Emissão'},
                    {'value': 'data_pagamento', 'label': 'Pagamento'},
                    {'value': 'valor', 'label': 'Valor'},
                    {'value': 'valor_pago', 'label': 'Valor Pago'},
                    {'value': 'status', 'label': 'Situação'},
                    {'value': 'tipo', 'label': 'Tipo'},
                ],
            },
        })


class RelatorioTitulosView(APIView):
    """
    GET /api/v1/relatorios/titulos/

    Query params:
        fonte        receber | pagar  (default: receber)
        formato      json | excel | csv  (default: json)
        campos       lista separada por vírgula (ex: participante_nome,valor_original,status)
        data_inicio  YYYY-MM-DD
        data_fim     YYYY-MM-DD
        data_field   data_vencimento | data_emissao | data_pagamento  (default: data_vencimento)
        status       aberto | baixado
        tipo         tipo do título
        busca        texto livre (filtra na descrição)
        ordenar_por  campo para ordenação
        ordenar_dir  asc | desc  (default: asc)
        page         número da página (somente para formato json, default: 1)
    """
    permission_classes = [IsAuthenticated]

    _CAMPOS_RECEBER = {
        'participante_nome': 'Participante',
        'tipo': 'Categoria',
        'descricao': 'Descrição',
        'data_emissao': 'Data de Emissão',
        'data_vencimento': 'Vencimento',
        'data_pagamento': 'Data de Pagamento',
        'valor_original': 'Valor (R$)',
        'valor_pago': 'Valor Pago (R$)',
        'multa': 'Multa (R$)',
        'saldo_devedor': 'Saldo Devedor (R$)',
        'status': 'Situação',
        'num_parcela': 'Parcela',
        'total_parcelas': 'Total Parcelas',
    }

    _CAMPOS_PAGAR = {
        'favorecido_nome': 'Favorecido',
        'tipo': 'Tipo',
        'descricao': 'Descrição',
        'data_emissao': 'Data de Emissão',
        'data_vencimento': 'Vencimento',
        'data_pagamento': 'Data de Pagamento',
        'valor': 'Valor (R$)',
        'valor_pago': 'Valor Pago (R$)',
        'status': 'Situação',
        'num_parcela': 'Parcela',
        'total_parcelas': 'Total Parcelas',
    }

    _SAFE_SORT_RECEBER = {
        'data_emissao', 'data_vencimento', 'data_pagamento',
        'valor_original', 'valor_pago', 'status', 'tipo',
        'num_parcela', 'descricao',
    }

    _SAFE_SORT_PAGAR = {
        'data_emissao', 'data_vencimento', 'data_pagamento',
        'valor', 'valor_pago', 'status', 'tipo',
        'num_parcela', 'descricao',
    }

    # ── Extratores de valores ────────────────────────────────────────────────

    def _fmt_date(self, d):
        return d.strftime('%d/%m/%Y') if d else ''

    def _fmt_decimal(self, v):
        return float(v) if v is not None else ''

    def _val_receber(self, obj, campo):
        try:
            if campo == 'participante_nome':
                if obj.participante:
                    return obj.participante.nome
                if obj.cliente:
                    return obj.cliente.nome
                return '—'
            if campo == 'tipo':
                return obj.get_tipo_display()
            if campo == 'status':
                return obj.get_status_display()
            if campo == 'saldo_devedor':
                return self._fmt_decimal(obj.saldo_devedor)
            if campo in ('valor_original', 'multa', 'valor_pago'):
                return self._fmt_decimal(getattr(obj, campo))
            if campo in ('data_emissao', 'data_vencimento', 'data_pagamento'):
                return self._fmt_date(getattr(obj, campo))
            return getattr(obj, campo, '')
        except Exception:
            return ''

    def _val_pagar(self, obj, campo):
        try:
            if campo == 'favorecido_nome':
                f = obj.favorecido
                if f.usuario:
                    return f.usuario.nome
                if f.entidade:
                    return f.entidade.nome
                return '—'
            if campo == 'tipo':
                return obj.get_tipo_display()
            if campo == 'status':
                return obj.get_status_display()
            if campo in ('valor', 'valor_pago'):
                return self._fmt_decimal(getattr(obj, campo))
            if campo in ('data_emissao', 'data_vencimento', 'data_pagamento'):
                return self._fmt_date(getattr(obj, campo))
            return getattr(obj, campo, '')
        except Exception:
            return ''

    # ── Queryset builders ───────────────────────────────────────────────────

    def _build_qs_receber(self, params):
        qs = (
            TituloReceber.objects
            .select_related('participante', 'cliente')
        )
        data_field = params.get('data_field', 'data_vencimento')
        if data_field not in ('data_emissao', 'data_vencimento', 'data_pagamento'):
            data_field = 'data_vencimento'

        if params.get('data_inicio'):
            qs = qs.filter(**{f'{data_field}__gte': params['data_inicio']})
        if params.get('data_fim'):
            qs = qs.filter(**{f'{data_field}__lte': params['data_fim']})
        if params.get('status'):
            qs = qs.filter(status=params['status'])
        if params.get('tipo'):
            qs = qs.filter(tipo=params['tipo'])
        if params.get('busca'):
            qs = qs.filter(descricao__icontains=params['busca'])

        sort_field = params.get('ordenar_por', 'data_vencimento')
        if sort_field not in self._SAFE_SORT_RECEBER:
            sort_field = 'data_vencimento'
        prefix = '-' if params.get('ordenar_dir', 'asc') == 'desc' else ''
        return qs.order_by(f'{prefix}{sort_field}')

    def _build_qs_pagar(self, params):
        qs = TituloPagar.objects.select_related(
            'favorecido__usuario', 'favorecido__entidade'
        )
        data_field = params.get('data_field', 'data_vencimento')
        if data_field not in ('data_emissao', 'data_vencimento', 'data_pagamento'):
            data_field = 'data_vencimento'

        if params.get('data_inicio'):
            qs = qs.filter(**{f'{data_field}__gte': params['data_inicio']})
        if params.get('data_fim'):
            qs = qs.filter(**{f'{data_field}__lte': params['data_fim']})
        if params.get('status'):
            qs = qs.filter(status=params['status'])
        if params.get('tipo'):
            qs = qs.filter(tipo=params['tipo'])
        if params.get('busca'):
            qs = qs.filter(descricao__icontains=params['busca'])

        sort_field = params.get('ordenar_por', 'data_vencimento')
        if sort_field not in self._SAFE_SORT_PAGAR:
            sort_field = 'data_vencimento'
        prefix = '-' if params.get('ordenar_dir', 'asc') == 'desc' else ''
        return qs.order_by(f'{prefix}{sort_field}')

    # ── Formatos de saída ───────────────────────────────────────────────────

    def _as_json(self, qs, campos, campos_config, extrator, page):
        page_size = 50
        total = qs.count()
        offset = (page - 1) * page_size
        rows = [
            {c: extrator(obj, c) for c in campos}
            for obj in qs[offset:offset + page_size]
        ]
        return Response({
            'count': total,
            'page': page,
            'page_size': page_size,
            'num_pages': max(1, (total + page_size - 1) // page_size),
            'campos': {c: campos_config[c] for c in campos},
            'results': rows,
        })

    def _as_excel(self, qs, campos, campos_config, extrator, fonte):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Títulos a {'Receber' if fonte == 'receber' else 'Pagar'}"

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')

        for col_i, campo in enumerate(campos, 1):
            cell = ws.cell(row=1, column=col_i, value=campos_config[campo])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_i, obj in enumerate(qs, 2):
            for col_i, campo in enumerate(campos, 1):
                ws.cell(row=row_i, column=col_i, value=extrator(obj, campo))

        # Auto-width
        for col_i, campo in enumerate(campos, 1):
            col_letter = get_column_letter(col_i)
            header_len = len(campos_config[campo])
            data_len = max(
                (len(str(ws.cell(row=r, column=col_i).value or ''))
                 for r in range(2, ws.max_row + 1)),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max(header_len, data_len) + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ts = timezone.now().strftime('%Y%m%d_%H%M')
        filename = f"relatorio_{fonte}_{ts}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    def _as_csv(self, qs, campos, campos_config, extrator, fonte):
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_ALL)
        writer.writerow([campos_config[c] for c in campos])
        for obj in qs:
            writer.writerow([extrator(obj, c) for c in campos])

        ts = timezone.now().strftime('%Y%m%d_%H%M')
        filename = f"relatorio_{fonte}_{ts}.csv"
        resp = HttpResponse(
            '﻿' + output.getvalue(),   # BOM UTF-8 para compatibilidade com Excel
            content_type='text/csv; charset=utf-8-sig',
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    # ── Handler principal ───────────────────────────────────────────────────

    def get(self, request):
        params = {k: v for k, v in request.query_params.items()}
        fonte = params.get('fonte', 'receber')
        formato = params.get('formato', 'json')
        campos_str = params.get('campos', '')
        page = max(1, int(params.get('page', 1)))

        if fonte == 'pagar':
            qs = self._build_qs_pagar(params)
            campos_config = self._CAMPOS_PAGAR
            extrator = self._val_pagar
        else:
            qs = self._build_qs_receber(params)
            campos_config = self._CAMPOS_RECEBER
            extrator = self._val_receber

        # Determina colunas solicitadas
        if campos_str:
            campos = [c for c in campos_str.split(',') if c in campos_config]
        if not campos_str or not campos:
            campos = list(campos_config.keys())

        if formato == 'excel':
            return self._as_excel(qs, campos, campos_config, extrator, fonte)
        if formato == 'csv':
            return self._as_csv(qs, campos, campos_config, extrator, fonte)
        return self._as_json(qs, campos, campos_config, extrator, page)
